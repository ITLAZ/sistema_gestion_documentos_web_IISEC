import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from difflib import SequenceMatcher

# Definir las rutas
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../'))
revistas_path = os.path.join(base_path, 'temp', 'ARTÍCULOS_ACADÉMICOS')
excel_path = os.path.join(os.path.dirname(__file__), 'BASE_DE_DATOS_PUBLICACIONES_ IISEC.xlsx')

# Verificar si la carpeta ARTÍCULOS_ACADÉMICOS existe
if not os.path.exists(revistas_path):
    raise FileNotFoundError(f"No se encontró la carpeta ARTÍCULOS_ACADÉMICOS en la ruta: {revistas_path}")

# Obtener los archivos de la carpeta ARTÍCULOS_ACADÉMICOS
revistas_files = [{'File Name': file.lower(), 'Path': os.path.join(revistas_path, file)} for file in os.listdir(revistas_path) if os.path.isfile(os.path.join(revistas_path, file))]

# Crear un DataFrame con las rutas de la carpeta ARTÍCULOS_ACADÉMICOS
revistas_df = pd.DataFrame(revistas_files)

# Leer el archivo Excel con encabezados en la tercera fila
excel_data = pd.read_excel(excel_path, sheet_name='Artículos en revistas', header=2)

# Función para extraer el número del artículo y el año
def extract_number_and_year(row):
    try:
        # Extraer contenido de la columna D
        content = row['Nombre de la Revista']
        if pd.isna(content):
            return None, None  # Si está vacío, saltar

        # Buscar "No." o "Edición Especial" para determinar prefijo
        if "No." in content:
            number = content.split("No.")[1].split(",")[0].strip()
        elif "Ed." or "Edición Especial" in content:
            number = "Especial"  # Para ediciones especiales
        else:
            number = None

        # Buscar el año en la columna correspondiente
        year_raw = row['Año de la revista']
        if pd.isna(year_raw):
            return number, None  # Si el año está vacío

        if isinstance(year_raw, datetime):
            year = str(year_raw.year)[-2:]  # Extraer los últimos dos dígitos si es formato datetime
        else:
            year = str(year_raw).strip()[-2:]  # Si es texto, tomar los últimos dos caracteres
        return number, year
    except Exception as e:
        print(f"Error al procesar fila: {row.name} -> {e}")
        return None, None

# Función para buscar el archivo por código
def find_file_path(number, year, used_files):
    if not number or not year:
        return None  # Si faltan datos, no buscar

    # Determinar prefijo según el número
    prefix = f"ee_{year}" if number == "Especial" else f"n{number}_{year}"

    # Buscar coincidencias exactas o con sufijos (_1, _2, _3)
    for i in range(1, 10):  # Aumentamos el rango por seguridad
        full_code = f"{prefix}_{i}.pdf"
        match = revistas_df[revistas_df['File Name'].str.contains(full_code)]
        if not match.empty:
            file_path = match.iloc[0]['Path']
            if file_path not in used_files:  # Verificar que no se haya usado ya
                used_files.add(file_path)
                return file_path

    # Caso especial sin sufijos
    for file in revistas_files:
        if file['File Name'].startswith(prefix):
            file_path = file['Path']
            if file_path not in used_files:  # Verificar que no se haya usado ya
                used_files.add(file_path)
                return file_path

    return None  # No se encontró archivo

# Función para buscar el archivo por comparación difusa
def find_file_by_similarity(row, used_files, threshold=0.4):
    # Las columnas G (índice 6) y H (índice 7) contienen los enlaces
    link_g = str(row[6]).lower() if not pd.isna(row[6]) else ""
    link_h = str(row[7]).lower() if not pd.isna(row[7]) else ""
    # Columnas de título y revista
    title = str(row['Título del artículo']).lower() if not pd.isna(row['Título del artículo']) else ""
    revista = str(row['Nombre de la Revista']).lower() if not pd.isna(row['Nombre de la Revista']) else ""

    best_match = None
    best_score = 0

    # Comparar primero los enlaces de las columnas G y H
    for _, file_row in revistas_df.iterrows():
        file_name = file_row['File Name']
        score_g = SequenceMatcher(None, link_g, file_name).ratio()
        score_h = SequenceMatcher(None, link_h, file_name).ratio()
        score = max(score_g, score_h)  # Tomar el mayor puntaje de los enlaces
        if score > best_score and score >= threshold and file_row['Path'] not in used_files:
            best_match = file_row['Path']
            best_score = score

    # Si no hay coincidencia con enlaces, comparar título y revista
    if not best_match:
        for _, file_row in revistas_df.iterrows():
            file_name = file_row['File Name']
            score_title = SequenceMatcher(None, title, file_name).ratio()
            score_revista = SequenceMatcher(None, revista, file_name).ratio()
            score = max(score_title, score_revista)  # Comparar título y nombre de la revista
            if score > best_score and score >= threshold and file_row['Path'] not in used_files:
                best_match = file_row['Path']
                best_score = score

    # Si hay coincidencia, marcar como usada
    if best_match:
        used_files.add(best_match)
    return best_match


# Realizar el mapeo de rutas
used_files = set()
mapped_paths = []
for _, row in excel_data.iterrows():
    number, year = extract_number_and_year(row)
    path = find_file_path(number, year, used_files) or find_file_by_similarity(row, used_files)
    mapped_paths.append(path)

# Cargar el archivo Excel existente usando openpyxl
workbook = load_workbook(excel_path)
sheet = workbook['Artículos en revistas']

# Colocar la columna PATH en la columna H
path_column_index = 9  # Columna I
if sheet.cell(row=3, column=path_column_index).value != 'PATH':
    sheet.cell(row=3, column=path_column_index, value='PATH')

# Añadir los valores de la columna PATH a partir de la fila 4
start_row = 4  # Fila donde empiezan los datos
to_update = zip(range(start_row, start_row + len(mapped_paths)), mapped_paths)
for i, path in to_update:
    sheet.cell(row=i, column=path_column_index, value=path)

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print(f"El archivo Excel existente se actualizó con la columna PATH en la columna I: {excel_path}")
