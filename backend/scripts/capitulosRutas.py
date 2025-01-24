import os
import pandas as pd
from openpyxl import load_workbook

# Definir las rutas
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../'))
capitulos_path = os.path.join(base_path, 'temp', 'CAPÍTULOS')
excel_path = os.path.join(os.path.dirname(__file__), 'BASE_DE_DATOS_PUBLICACIONES_ IISEC.xlsx')

# Verificar si la carpeta CAPÍTULOS existe
if not os.path.exists(capitulos_path):
    raise FileNotFoundError(f"No se encontró la carpeta CAPÍTULOS en la ruta: {capitulos_path}")

# Obtener los archivos de la carpeta CAPÍTULOS
capitulos_files = [{'File Name': file.lower(), 'Path': os.path.join(capitulos_path, file)} for file in os.listdir(capitulos_path) if os.path.isfile(os.path.join(capitulos_path, file))]

# Crear un DataFrame con las rutas de la carpeta CAPÍTULOS
capitulos_df = pd.DataFrame(capitulos_files)

# Leer el archivo Excel con encabezados en la tercera fila
excel_data = pd.read_excel(excel_path, sheet_name='Capítulos de libros', header=2)

# Función para buscar el archivo por comparación exacta con el enlace de la columna I
def find_file_by_link(row, used_files):
    # Extraer el link de la columna I
    link = str(row[8]).strip().lower() if not pd.isna(row[8]) else ""  # Columna I sin nombre

    # Extraer el identificador único del archivo desde el link
    link_identifier = os.path.basename(link) if link else ""  # Extraer el nombre del archivo del enlace

    for _, file_row in capitulos_df.iterrows():
        file_name = file_row['File Name']
        # Comparar el identificador del link con el nombre del archivo
        if link_identifier == file_name and file_row['Path'] not in used_files:
            used_files.add(file_row['Path'])
            return file_row['Path']

    return None

# Realizar el mapeo de rutas
used_files = set()
mapped_paths = []
for _, row in excel_data.iterrows():
    path = find_file_by_link(row, used_files)
    mapped_paths.append(path)

# Cargar el archivo Excel existente usando openpyxl
workbook = load_workbook(excel_path)
sheet = workbook['Capítulos de libros']

# Colocar la columna PATH en la columna L
path_column_index = 12  # Columna L
if sheet.cell(row=3, column=path_column_index).value != 'PATH':
    sheet.cell(row=3, column=path_column_index, value='PATH')

# Añadir los valores de la columna PATH a partir de la fila 4
start_row = 4  # Fila donde empiezan los datos
to_update = zip(range(start_row, start_row + len(mapped_paths)), mapped_paths)
for i, path in to_update:
    sheet.cell(row=i, column=path_column_index, value=path)

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print(f"El archivo Excel existente se actualizó con la columna PATH en la columna L: {excel_path}")
