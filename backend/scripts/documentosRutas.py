import os
import pandas as pd
from openpyxl import load_workbook
from difflib import SequenceMatcher

# Definir las rutas
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../'))
documentos_path = os.path.join(base_path, 'temp', 'DOCUMENTOS_DE_TRABAJO')
excel_path = os.path.join(os.path.dirname(__file__), 'BASE_DE_DATOS_PUBLICACIONES_ IISEC.xlsx')

# Verificar si la carpeta DOCUMENTOS_DE_TRABAJO existe
if not os.path.exists(documentos_path):
    raise FileNotFoundError(f"No se encontró la carpeta DOCUMENTOS_DE_TRABAJO en la ruta: {documentos_path}")

# Obtener los archivos de la carpeta DOCUMENTOS_DE_TRABAJO
documentos_files = [{'File Name': file.lower(), 'Path': os.path.join(documentos_path, file)}
                    for file in os.listdir(documentos_path) if os.path.isfile(os.path.join(documentos_path, file))]

# Crear un DataFrame con las rutas de la carpeta DOCUMENTOS_DE_TRABAJO
documentos_df = pd.DataFrame(documentos_files)

# Leer el archivo Excel con encabezados en la tercera fila
excel_data = pd.read_excel(excel_path, sheet_name='Documentos de Trabajo', header=2)

# Función para buscar el archivo por comparación difusa con el enlace (columna G)
def find_file_by_link(row, used_files, threshold=0.1):
    # Obtener el enlace de la columna G
    link = str(row['Link PDF']).lower() if not pd.isna(row['Link PDF']) else ""

    best_match = None
    best_score = 0

    # Comparar los enlaces con los nombres de archivo
    for _, file_row in documentos_df.iterrows():
        file_name = file_row['File Name']
        score = SequenceMatcher(None, link, file_name).ratio()
        if score > best_score and score >= threshold and file_row['Path'] not in used_files:
            best_match = file_row['Path']
            best_score = score

    # Marcar como usado si se encontró
    if best_match:
        used_files.add(best_match)
    return best_match

# Realizar el mapeo de rutas
used_files = set()
mapped_paths = []

for _, row in excel_data.iterrows():
    path = find_file_by_link(row, used_files)
    mapped_paths.append(path)

# Cargar el archivo Excel existente usando openpyxl
workbook = load_workbook(excel_path)
sheet = workbook['Documentos de Trabajo']

# Colocar la columna PATH en la columna J
path_column_index = 10  # Columna J
if sheet.cell(row=3, column=path_column_index).value != 'PATH':
    sheet.cell(row=3, column=path_column_index, value='PATH')

# Añadir los valores de la columna PATH a partir de la fila 4
start_row = 4  # Fila donde empiezan los datos
to_update = zip(range(start_row, start_row + len(mapped_paths)), mapped_paths)
for i, path in to_update:
    sheet.cell(row=i, column=path_column_index, value=path)

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print(f"El archivo Excel existente se actualizó con la columna PATH en la columna J: {excel_path}")
