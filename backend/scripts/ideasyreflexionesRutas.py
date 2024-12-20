import os
import pandas as pd
from openpyxl import load_workbook
from difflib import SequenceMatcher

# Definir las rutas
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../'))
ideas_path = os.path.join(base_path, 'temp', 'IDEAS_Y_REFLEXIONES')
excel_path = os.path.join(os.path.dirname(__file__), 'BASE_DE_DATOS_PUBLICACIONES_ IISEC.xlsx')

# Verificar si la carpeta IDEAS_Y_REFLEXIONES existe
if not os.path.exists(ideas_path):
    raise FileNotFoundError(f"No se encontró la carpeta IDEAS_Y_REFLEXIONES en la ruta: {ideas_path}")

# Obtener los archivos de la carpeta IDEAS_Y_REFLEXIONES
ideas_files = [{'File Name': file.lower(), 'Path': os.path.join(ideas_path, file)} for file in os.listdir(ideas_path) if os.path.isfile(os.path.join(ideas_path, file))]

# Crear un DataFrame con las rutas de la carpeta IDEAS_Y_REFLEXIONES
ideas_df = pd.DataFrame(ideas_files)

# Leer el archivo Excel con encabezados en la tercera fila
excel_data = pd.read_excel(excel_path, sheet_name='Ideas y Reflexiones', header=2)

# Función para buscar el archivo por comparación exacta o difusa con la columna de enlaces
def find_file_by_link(row, used_files, threshold=0.1):
    link = str(row['Link PDF']).strip().lower() if not pd.isna(row['Link PDF']) else ""

    best_match = None
    best_score = 0

    for _, file_row in ideas_df.iterrows():
        file_name = file_row['File Name']
        score = SequenceMatcher(None, link, file_name).ratio()  # Comparar usando similaridad
        if score > best_score and score >= threshold and file_row['Path'] not in used_files:
            best_match = file_row['Path']
            best_score = score

    if best_match:
        used_files.add(best_match)  # Marcar como usado
    return best_match

# Realizar el mapeo de rutas
used_files = set()
mapped_paths = []
for _, row in excel_data.iterrows():
    path = find_file_by_link(row, used_files)
    mapped_paths.append(path)

# Cargar el archivo Excel existente usando openpyxl
workbook = load_workbook(excel_path)
sheet = workbook['Ideas y Reflexiones']

# Colocar la columna PATH en la columna G
path_column_index = 7  # Columna G
if sheet.cell(row=3, column=path_column_index).value != 'PATH':
    sheet.cell(row=3, column=path_column_index, value='PATH')

# Añadir los valores de la columna PATH a partir de la fila 4
start_row = 4  # Fila donde empiezan los datos
to_update = zip(range(start_row, start_row + len(mapped_paths)), mapped_paths)
for i, path in to_update:
    sheet.cell(row=i, column=path_column_index, value=path)

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print(f"El archivo Excel existente se actualizó con la columna PATH en la columna G: {excel_path}")
