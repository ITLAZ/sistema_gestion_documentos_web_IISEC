import os
import difflib
import pandas as pd
from openpyxl import load_workbook

# Definir las rutas
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../'))
info_iisec_path = os.path.join(base_path, 'temp', 'INFO_IISEC')
excel_path = os.path.join(os.path.dirname(__file__), 'BASE_DE_DATOS_PUBLICACIONES_ IISEC.xlsx')

# Verificar si la carpeta INFO_IISEC existe
if not os.path.exists(info_iisec_path):
    raise FileNotFoundError(f"No se encontró la carpeta INFO_IISEC en la ruta: {info_iisec_path}")

# Obtener los archivos de la carpeta INFO_IISEC
info_files = [{'File Name': file.lower(), 'Path': os.path.join(info_iisec_path, file)} for file in os.listdir(info_iisec_path) if os.path.isfile(os.path.join(info_iisec_path, file))]

# Crear un DataFrame con las rutas de la carpeta INFO_IISEC
info_df = pd.DataFrame(info_files)

# Leer el archivo Excel con encabezados en la tercera fila
excel_data = pd.read_excel(excel_path, sheet_name='Info IISEC', header=2)

# Función para buscar el archivo más similar al link usando SequenceMatcher
def find_file_by_link(row, used_files, threshold=0.2):
    link = str(row['Link PDF']).lower().strip() if not pd.isna(row['Link PDF']) else ""  # Columna F

    best_match = None
    highest_ratio = 0
    for _, file_row in info_df.iterrows():
        file_name = file_row['File Name']
        similarity = difflib.SequenceMatcher(None, link, file_name).ratio()
        if similarity > highest_ratio and similarity >= threshold and file_row['Path'] not in used_files:
            highest_ratio = similarity
            best_match = file_row['Path']

    if best_match:
        used_files.add(best_match)
        return best_match
    return None

# Realizar el mapeo de rutas
used_files = set()
mapped_paths = []
for _, row in excel_data.iterrows():
    path = find_file_by_link(row, used_files)
    mapped_paths.append(path)

# Cargar el archivo Excel existente usando openpyxl
workbook = load_workbook(excel_path)
sheet = workbook['Info IISEC']

# Colocar la columna PATH en la columna G
path_column_index = 7  # Columna G
if sheet.cell(row=3, column=path_column_index).value != 'PATH':
    sheet.cell(row=3, column=path_column_index, value='PATH')

# Añadir los valores de la columna PATH a partir de la fila 4
start_row = 4  # Fila donde empiezan los datos
to_update = zip(range(start_row, start_row + len(mapped_paths)), mapped_paths)
for i, path in to_update:
    sheet.cell(row=i, column=path_column_index, value=path)

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print(f"El archivo Excel existente se actualizó con la columna PATH en la columna G: {excel_path}")
