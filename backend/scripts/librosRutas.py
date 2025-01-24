import os
import pandas as pd
from openpyxl import load_workbook
from difflib import SequenceMatcher

# Definir las rutas
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../'))
libros_path = os.path.join(base_path, 'temp', 'LIBROS')  # Carpeta LIBROS
excel_path = os.path.join(os.path.dirname(__file__), 'BASE_DE_DATOS_PUBLICACIONES_ IISEC.xlsx')

# Obtener los archivos de la carpeta LIBROS
libros_files = [f for f in os.listdir(libros_path) if os.path.isfile(os.path.join(libros_path, f))]
libros_files = [{'File Name': file.lower(), 'Path': os.path.join(libros_path, file)} for file in libros_files]

# Crear un DataFrame con las rutas de la carpeta LIBROS
libros_df = pd.DataFrame(libros_files)

# Leer el archivo Excel con encabezados en la tercera fila
excel_data = pd.read_excel(excel_path, sheet_name='Libros', header=2)

# Normalizar nombres en el Excel para facilitar la comparación
excel_data['Título del libro'] = excel_data['Título del libro'].str.lower().str.strip()

# Función para encontrar la mejor coincidencia
def find_best_match(title, files, threshold=0.2):
    best_match = None
    best_score = 0
    for file in files:
        score = SequenceMatcher(None, title, file['File Name']).ratio()
        if score > best_score and score >= threshold:
            best_match = file
            best_score = score
    return best_match

# Realizar el mapeo de rutas
mapped_paths = []
for _, row in excel_data.iterrows():
    title = row['Título del libro']
    match = find_best_match(title, libros_files)
    if match:
        mapped_paths.append(match['Path'])
    else:
        mapped_paths.append(None)

# Cargar el archivo Excel existente usando openpyxl
workbook = load_workbook(excel_path)
sheet = workbook['Libros']

# Colocar la columna PATH en la columna H
path_column_index = 8  # Columna H
if sheet.cell(row=3, column=path_column_index).value != 'PATH':
    sheet.cell(row=3, column=path_column_index, value='PATH')

# Añadir los valores de la columna PATH a partir de la fila 4
start_row = 4  # Fila donde empiezan los datos
for i, path in enumerate(mapped_paths, start=start_row):
    sheet.cell(row=i, column=path_column_index, value=path)

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print(f"El archivo Excel existente se actualizó con la columna PATH en la columna H: {excel_path}")
