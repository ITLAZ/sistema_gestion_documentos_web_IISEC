import os
import pandas as pd
from openpyxl import load_workbook

# Definir las rutas
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../'))
policy_brief_path = os.path.join(base_path, 'temp', 'POLICY_BRIEF')
excel_path = os.path.join(os.path.dirname(__file__), 'BASE_DE_DATOS_PUBLICACIONES_ IISEC.xlsx')

# Verificar si la carpeta POLICY_BRIEF existe
if not os.path.exists(policy_brief_path):
    raise FileNotFoundError(f"No se encontró la carpeta POLICY_BRIEF en la ruta: {policy_brief_path}")

# Obtener los archivos de la carpeta POLICY_BRIEF
policy_files = [{'File Name': file.lower(), 'Path': os.path.join(policy_brief_path, file)} for file in os.listdir(policy_brief_path) if os.path.isfile(os.path.join(policy_brief_path, file))]

# Crear un DataFrame con las rutas de la carpeta POLICY_BRIEF
policy_df = pd.DataFrame(policy_files)

# Leer el archivo Excel con encabezados en la tercera fila
excel_data = pd.read_excel(excel_path, sheet_name='Policy Brief', header=2)

# Función para buscar el archivo basado en el número en la columna N°
def find_file_by_number(row, used_files):
    number = str(row[0])  # Columna N° (columna A)
    file_name_pattern = f"policy-brief-iisec-{number}.pdf"

    for _, file_row in policy_df.iterrows():
        file_name = file_row['File Name']
        if file_name == file_name_pattern.lower() and file_row['Path'] not in used_files:
            used_files.add(file_row['Path'])
            return file_row['Path']

    return None

# Realizar el mapeo de rutas
used_files = set()
mapped_paths = []
for _, row in excel_data.iterrows():
    path = find_file_by_number(row, used_files)
    mapped_paths.append(path)

# Cargar el archivo Excel existente usando openpyxl
workbook = load_workbook(excel_path)
sheet = workbook['Policy Brief']

# Colocar la columna PATH en la columna G
path_column_index = 7  # Columna G
if sheet.cell(row=3, column=path_column_index).value != 'PATH':
    sheet.cell(row=3, column=path_column_index, value='PATH')

# Añadir los valores de la columna PATH a partir de la fila 4
start_row = 4  # Fila donde empiezan los datos
to_update = zip(range(start_row, start_row + len(mapped_paths)), mapped_paths)
for i, path in to_update:
    sheet.cell(row=i, column=path_column_index, value=path)

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print(f"El archivo Excel existente se actualizó con la columna PATH en la columna G: {excel_path}")
